import csv
import io
import zipfile

from openpyxl import Workbook

from apps.accounts.models import ParticipantProfile
from apps.survey.comment_identity import comment_display_name
from apps.survey.models import CommentReaction, ConversationMessage, EnglishPaperResponse, PostReaction, QualityEvent, ScaleResponse, SurveySession, TextResponse, TopicRound


SECTIONS = {
    "participants": "participants",
    "snapshots": "snapshots",
    "topic_order": "topic_order",
    "rounds": "rounds",
    "comment_reactions": "comment_reactions",
    "post_reactions": "post_reactions",
    "english_papers": "english_papers",
    "scale_responses": "scale_responses",
    "text_responses": "text_responses",
    "conversation": "conversation",
    "quality_events": "quality_events",
}


def participant_rows(batch):
    return [
        {
            "username": profile.user.username,
            "display_name": profile.display_name,
            "region": profile.region,
            "gender": profile.gender,
            "contact": profile.contact,
        }
        for profile in batch.participants.select_related("user")
    ]


def snapshot_rows(batch):
    return [{"username": session.user.username, "batch_snapshot": session.batch_snapshot} for session in batch.sessions.select_related("user")]


def topic_order_rows(batch):
    return [
        {
            "username": session.user.username,
            "initial_order": session.topic_order_snapshot,
            "submitted_order": session.submitted_topic_order,
            "high_topic_id": session.selected_high_topic_id,
            "low_topic_id": session.selected_low_topic_id,
        }
        for session in batch.sessions.select_related("user")
    ]


def round_rows(batch):
    return [
        {
            "username": round_obj.session.user.username,
            "round_type": round_obj.round_type,
            "topic_id": round_obj.topic_id,
            "skipped_ai": round_obj.skipped_ai,
            "completed": round_obj.is_completed,
        }
        for round_obj in _rounds(batch)
    ]


def comment_reaction_rows(batch):
    return [
        {
            "username": reaction.round.session.user.username,
            "round_type": reaction.round.round_type,
            "comment_snapshot_id": reaction.comment_snapshot_id,
            "reaction": reaction.reaction,
        }
        for reaction in _reaction_queryset(batch)
    ]


def post_reaction_rows(batch):
    return [
        {
            "username": reaction.round.session.user.username,
            "round_type": reaction.round.round_type,
            "topic_id": reaction.round.topic_id,
            "reaction": reaction.reaction,
        }
        for reaction in PostReaction.objects.filter(round__session__batch=batch).select_related("round__session__user")
    ]


def english_paper_rows(batch):
    return [
        {
            "username": response.session.user.username,
            "prompt": response.prompt,
            "duration_hours": response.duration_hours,
            "paper_text": response.paper_text,
            "submitted_at": response.submitted_at.isoformat(),
        }
        for response in EnglishPaperResponse.objects.filter(session__batch=batch).select_related("session__user")
    ]


def scale_response_rows(batch):
    return [
        {
            "username": response.round.session.user.username,
            "step": response.step,
            "item_type": response.item_type,
            "item_label": response.item_label,
            "selected_value": response.selected_value,
        }
        for response in ScaleResponse.objects.filter(round__session__batch=batch).select_related("round__session__user")
    ]


def text_response_rows(batch):
    return [
        {
            "username": response.round.session.user.username,
            "step": response.step,
            "final_text": response.final_text,
            "input_method": response.input_method,
            "word_count": response.word_count,
        }
        for response in TextResponse.objects.filter(round__session__batch=batch).select_related("round__session__user")
    ]


def conversation_rows(batch):
    return [
        {
            "username": message.round.session.user.username,
            "round_type": message.round.round_type,
            "role": message.role,
            "content": message.content,
            "model_name": message.model_name,
        }
        for message in ConversationMessage.objects.filter(round__session__batch=batch).select_related("round__session__user")
    ]


def quality_event_rows(batch):
    return [
        {
            "username": event.user.username,
            "event_type": event.event_type,
            "metadata": event.metadata,
            "created_at": event.created_at.isoformat(),
        }
        for event in QualityEvent.objects.filter(session__batch=batch).select_related("user")
    ]


SECTION_BUILDERS = {
    "participants": participant_rows,
    "snapshots": snapshot_rows,
    "topic_order": topic_order_rows,
    "rounds": round_rows,
    "comment_reactions": comment_reaction_rows,
    "post_reactions": post_reaction_rows,
    "english_papers": english_paper_rows,
    "scale_responses": scale_response_rows,
    "text_responses": text_response_rows,
    "conversation": conversation_rows,
    "quality_events": quality_event_rows,
}


def _rounds(batch):
    return TopicRound.objects.filter(session__batch=batch).select_related("session__user")


def _reaction_queryset(batch):
    from apps.survey.models import CommentReaction

    return CommentReaction.objects.filter(round__session__batch=batch).select_related("round__session__user")


def selected_rows(batch, sections):
    return {section: SECTION_BUILDERS[section](batch) for section in sections if section in SECTION_BUILDERS}


def build_excel(batch, sections):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for section, rows in selected_rows(batch, sections).items():
        sheet = workbook.create_sheet(section[:31])
        headers = sorted(rows[0].keys()) if rows else ["empty"]
        sheet.append(headers)
        for row in rows:
            sheet.append([str(row.get(header, "")) for header in headers])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def _joined_values(values):
    return "\n".join(str(value) for value in values if value not in [None, ""])


def _scale_summary(rounds, step):
    rows = ScaleResponse.objects.filter(round__in=rounds, step=step).order_by("submitted_at", "id")
    return _joined_values(f"{row.item_label}={row.selected_value}" for row in rows)


def _text_for_step(rounds, step):
    rows = TextResponse.objects.filter(round__in=rounds, step=step).order_by("submitted_at", "id")
    return _joined_values(row.final_text for row in rows)


def _conversation_text(rounds):
    messages = ConversationMessage.objects.filter(round__in=rounds).order_by("created_at", "id")
    return _joined_values(f"{message.get_role_display()}：{message.content}" for message in messages if message.content)


def _comment_by_snapshot_id(round_obj, comment_snapshot_id):
    for index, comment in enumerate(round_obj.material_snapshot.get("comments", [])):
        if comment.get("id") == comment_snapshot_id:
            return comment, index
    return {}, None


def _comment_reaction_label(reaction):
    labels = {"like": "赞", "dislike": "踩", "none": "无操作"}
    return labels.get(reaction, reaction)


SESSION_STEP_LABELS = {
    "research_consent": "参与研究授权同意",
    "topic_order": "话题排序",
    "round": "话题轮次",
    "english_paper": "英文论文写作",
    "done": "已完成",
}

ROUND_EXPORT_LABELS = {
    TopicRound.HIGH: "高分话题",
    TopicRound.LOW: "低分话题",
}

PROFILE_EXPORT_HEADERS = [
    "资料-用户名",
    "资料-称呼/姓名",
    "资料-地区",
    "资料-年龄段",
    "资料-性别",
    "资料-学校/单位类型",
    "资料-教育/职业状态",
    "资料-联系方式",
    "资料-备注",
]

STATUS_EXPORT_HEADERS = [
    "资料-用户名",
    "流程状态-当前步骤",
    "流程状态-轮次数",
    "流程状态-英文论文状态",
    "流程状态-开始时间",
    "流程状态-完成时间",
]

TOPIC_ORDER_EXPORT_HEADERS = ["资料-用户名", "话题排序"]

PAPER_EXPORT_HEADERS = [
    "资料-用户名",
    "英文论文-时长",
    "英文论文-提交时间",
    "英文论文-内容",
]


def _indexed_headers(rows, prefix):
    max_index = 0
    for row in rows or []:
        for key in row.keys():
            if key.startswith(prefix):
                suffix = key.removeprefix(prefix)
                if suffix.isdigit():
                    max_index = max(max_index, int(suffix))
    max_index = max(max_index, 1)
    return [f"{prefix}{index}" for index in range(1, max_index + 1)]


def _round_export_headers(round_label, rows=None):
    return [
        "资料-用户名",
        f"{round_label}-题目",
        *_indexed_headers(rows, f"{round_label}-量表记录"),
        *_indexed_headers(rows, f"{round_label}-回复内容"),
        *_indexed_headers(rows, f"{round_label}-帖子和评论互动"),
        *_indexed_headers(rows, f"{round_label}-AI对话"),
    ]


def _all_users_csv_headers(rows=None):
    headers = list(PROFILE_EXPORT_HEADERS)
    for group in [
        STATUS_EXPORT_HEADERS,
        TOPIC_ORDER_EXPORT_HEADERS,
        _round_export_headers("高分话题", rows),
        _round_export_headers("低分话题", rows),
        PAPER_EXPORT_HEADERS,
    ]:
        headers.extend(header for header in group if header != "资料-用户名")
    return headers


def _session_for_profile(profile):
    try:
        return profile.user.survey_session
    except SurveySession.DoesNotExist:
        return None


def _format_dt(value):
    return value.isoformat() if value else ""


def _english_paper_response(session):
    if not session:
        return None
    try:
        return session.english_paper_response
    except EnglishPaperResponse.DoesNotExist:
        return None


def _snapshot_topic_title(snapshot):
    return (
        snapshot.get("title_zh")
        or snapshot.get("title")
        or snapshot.get("title_en")
        or snapshot.get("statement_zh")
        or snapshot.get("statement")
        or snapshot.get("statement_en")
        or f"话题 {snapshot.get('id', '')}".strip()
    )


def _topic_title_for_id(session, topic_id):
    if session:
        for item in session.topic_order_snapshot or []:
            if str(item.get("id")) == str(topic_id):
                return _snapshot_topic_title(item)
    return f"话题 {topic_id}"


def _round_topic_title(round_obj, session):
    if not round_obj:
        return ""
    return (
        round_obj.material_snapshot.get("title_zh")
        or round_obj.material_snapshot.get("title")
        or round_obj.material_snapshot.get("title_en")
        or _topic_title_for_id(session, round_obj.topic_id)
    )


def _topic_order_text(session):
    if not session or not session.submitted_topic_order:
        return ""
    values = []
    for index, topic_id in enumerate(session.submitted_topic_order, start=1):
        marker = ""
        if int(topic_id) == session.selected_high_topic_id:
            marker = "（高分话题）"
        elif int(topic_id) == session.selected_low_topic_id:
            marker = "（低分话题）"
        values.append(f"{index}. {_topic_title_for_id(session, topic_id)}{marker}")
    return "\n".join(values)


def _ordered_rounds(session):
    if not session:
        return []
    order = {TopicRound.HIGH: 0, TopicRound.LOW: 1}
    return sorted(session.rounds.all(), key=lambda round_obj: (order.get(round_obj.round_type, 99), round_obj.pk))


def _round_for_type(rounds, round_type):
    return next((round_obj for round_obj in rounds if round_obj.round_type == round_type), None)


def _round_scale_values(round_obj):
    if not round_obj:
        return []
    rows = ScaleResponse.objects.filter(round=round_obj).order_by("submitted_at", "id")
    return [f"{row.item_label}={row.selected_value}" for row in rows]


def _round_reply_values(round_obj):
    if not round_obj:
        return []
    rows = TextResponse.objects.filter(round=round_obj).order_by("submitted_at", "id")
    return [row.final_text for row in rows]


def _round_interaction_values(round_obj, session):
    if not round_obj:
        return []
    values = []
    for row in PostReaction.objects.filter(round=round_obj).order_by("submitted_at", "id"):
        values.append(f"帖子: {_round_topic_title(round_obj, session)}: {_comment_reaction_label(row.reaction)}")
    for row in CommentReaction.objects.filter(round=round_obj).order_by("submitted_at", "id"):
        comment, comment_index = _comment_by_snapshot_id(round_obj, row.comment_snapshot_id)
        values.append(
            f"评论: 评论 #{row.comment_snapshot_id} ({comment_display_name(comment_index)}): "
            f"{_comment_reaction_label(row.reaction)}"
        )
    return values


def _round_conversation_values(round_obj):
    if not round_obj:
        return []
    rows = ConversationMessage.objects.filter(round=round_obj).order_by("created_at", "id")
    return [f"{row.get_role_display()}: {row.content}" for row in rows if row.content]


def _add_indexed_values(row, prefix, values):
    for index, value in enumerate(values, start=1):
        row[f"{prefix}{index}"] = value


def _profile_export_row(profile):
    return {
        "资料-用户名": profile.user.username,
        "资料-称呼/姓名": profile.display_name,
        "资料-地区": profile.region,
        "资料-年龄段": profile.age_range,
        "资料-性别": profile.gender,
        "资料-学校/单位类型": profile.organization_type,
        "资料-教育/职业状态": profile.education_or_work,
        "资料-联系方式": profile.contact,
        "资料-备注": profile.notes,
    }


def _status_export_row(profile, session, rounds):
    paper = _english_paper_response(session)
    return {
        "资料-用户名": profile.user.username,
        "流程状态-当前步骤": SESSION_STEP_LABELS.get(session.current_session_step, session.current_session_step) if session else "-",
        "流程状态-轮次数": len(rounds),
        "流程状态-英文论文状态": "已提交" if paper else "-",
        "流程状态-开始时间": _format_dt(session.started_at) if session else "",
        "流程状态-完成时间": _format_dt(session.completed_at) if session and session.completed_at else "",
    }


def _topic_order_export_row(profile, session):
    return {"资料-用户名": profile.user.username, "话题排序": _topic_order_text(session)}


def _round_export_row(profile, session, rounds, round_type):
    round_label = ROUND_EXPORT_LABELS[round_type]
    round_obj = _round_for_type(rounds, round_type)
    row = {
        "资料-用户名": profile.user.username,
        f"{round_label}-题目": _round_topic_title(round_obj, session),
    }
    _add_indexed_values(row, f"{round_label}-量表记录", _round_scale_values(round_obj))
    _add_indexed_values(row, f"{round_label}-回复内容", _round_reply_values(round_obj))
    _add_indexed_values(row, f"{round_label}-帖子和评论互动", _round_interaction_values(round_obj, session))
    _add_indexed_values(row, f"{round_label}-AI对话", _round_conversation_values(round_obj))
    return row


def _paper_export_row(profile, session):
    paper = _english_paper_response(session)
    return {
        "资料-用户名": profile.user.username,
        "英文论文-时长": paper.duration_hours if paper else "",
        "英文论文-提交时间": _format_dt(paper.submitted_at) if paper else "",
        "英文论文-内容": paper.paper_text if paper else "",
    }


def _export_context_rows(profile):
    session = _session_for_profile(profile)
    rounds = _ordered_rounds(session)
    return {
        "profile": _profile_export_row(profile),
        "status": _status_export_row(profile, session, rounds),
        "topic_order": _topic_order_export_row(profile, session),
        "high": _round_export_row(profile, session, rounds, TopicRound.HIGH),
        "low": _round_export_row(profile, session, rounds, TopicRound.LOW),
        "paper": _paper_export_row(profile, session),
    }


def _all_export_context_rows():
    profiles = (
        ParticipantProfile.objects.select_related("user", "batch")
        .filter(user__is_staff=False, user__is_superuser=False)
        .order_by("user__username")
    )
    return [_export_context_rows(profile) for profile in profiles]


def _combined_export_row(context_rows):
    row = {}
    for group in ["profile", "status", "topic_order", "high", "low", "paper"]:
        row.update(context_rows[group])
    return row


def _post_reactions_text(rounds):
    rows = PostReaction.objects.filter(round__in=rounds).select_related("round").order_by("submitted_at", "id")
    return _joined_values(
        f"{row.round.round_type}/topic-{row.round.topic_id}: {_comment_reaction_label(row.reaction)}"
        for row in rows
    )


def _comment_reactions_text(rounds):
    rows = CommentReaction.objects.filter(round__in=rounds).select_related("round").order_by("submitted_at", "id")
    values = []
    for row in rows:
        comment, comment_index = _comment_by_snapshot_id(row.round, row.comment_snapshot_id)
        values.append(
            f"{row.round.round_type}/topic-{row.round.topic_id}/comment-{row.comment_snapshot_id}"
            f" ({comment_display_name(comment_index)}): {_comment_reaction_label(row.reaction)}"
        )
    return _joined_values(values)


def all_users_summary_rows():
    return [_combined_export_row(context_rows) for context_rows in _all_export_context_rows()]


def build_all_users_csv():
    rows = all_users_summary_rows()
    headers = _all_users_csv_headers(rows)
    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return output.getvalue().encode("utf-8")


def build_all_users_excel():
    workbook = Workbook()
    workbook.remove(workbook.active)
    context_rows = _all_export_context_rows()
    sheet_specs = [
        ("资料", PROFILE_EXPORT_HEADERS, "profile"),
        ("流程状态", STATUS_EXPORT_HEADERS, "status"),
        ("话题排序", TOPIC_ORDER_EXPORT_HEADERS, "topic_order"),
        ("高分话题", _round_export_headers("高分话题", [row["high"] for row in context_rows]), "high"),
        ("低分话题", _round_export_headers("低分话题", [row["low"] for row in context_rows]), "low"),
        ("英文论文", PAPER_EXPORT_HEADERS, "paper"),
    ]
    sheets = {}
    for title, headers, key in sheet_specs:
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        sheets[key] = (sheet, headers)

    for context_row in context_rows:
        for key, (sheet, headers) in sheets.items():
            row = context_row[key]
            sheet.append([row.get(header, "") for header in headers])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def build_csv_zip(batch, sections):
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for section, rows in selected_rows(batch, sections).items():
            buffer = io.StringIO()
            headers = sorted(rows[0].keys()) if rows else ["empty"]
            writer = csv.DictWriter(buffer, fieldnames=headers)
            writer.writeheader()
            for row in rows:
                writer.writerow({header: row.get(header, "") for header in headers})
            archive.writestr(f"{section}.csv", buffer.getvalue())
    output.seek(0)
    return output.getvalue()
