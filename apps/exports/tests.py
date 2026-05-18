import csv
import zipfile
from io import BytesIO, StringIO

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from apps.experiments.models import ExperimentBatch
from apps.survey.models import CommentReaction, ConversationMessage, EnglishPaperResponse, PostReaction, QualityEvent, ScaleResponse, SurveySession, TextResponse, TopicRound

from .services import build_all_users_csv, build_all_users_excel, build_csv_zip, build_excel


class ExportServiceTests(TestCase):
    def setUp(self):
        self.batch = ExperimentBatch.objects.create(name="批次 A")
        user = User.objects.create_user("p001", password="pass")
        user.participant_profile.display_name = "参与者"
        user.participant_profile.batch = self.batch
        user.participant_profile.save()
        session = SurveySession.objects.create(user=user, batch=self.batch, topic_order_snapshot=[])
        round_obj = TopicRound.objects.create(session=session, round_type=TopicRound.HIGH, topic_id=1)
        ScaleResponse.objects.create(round=round_obj, step="emotion", item_type="emotion", item_label="放松", language="zh-hans", selected_value=5)
        ScaleResponse.objects.create(round=round_obj, step="ai_eval", item_type="ai", item_label="有帮助", language="zh-hans", selected_value=6)
        TextResponse.objects.create(round=round_obj, step="initial_text", final_text="想法", word_count=2)
        ConversationMessage.objects.create(round=round_obj, role="participant", content="你好", language="zh-hans")
        ConversationMessage.objects.create(round=round_obj, role="assistant", content="回应", language="zh-hans")
        PostReaction.objects.create(round=round_obj, reaction="like")
        CommentReaction.objects.create(round=round_obj, comment_snapshot_id=1, reaction="dislike")
        EnglishPaperResponse.objects.create(session=session, prompt="Write an essay.", duration_hours=24, paper_text="My English essay.")
        QualityEvent.objects.create(user=user, session=session, event_type="copy")

    def test_excel_contains_selected_sheets(self):
        payload = build_excel(self.batch, ["participants", "scale_responses"])

        workbook = load_workbook(BytesIO(payload))

        self.assertIn("participants", workbook.sheetnames)
        self.assertIn("scale_responses", workbook.sheetnames)

    def test_excel_can_export_post_reactions(self):
        payload = build_excel(self.batch, ["post_reactions"])

        workbook = load_workbook(BytesIO(payload))
        self.assertIn("post_reactions", workbook.sheetnames)
        sheet = workbook["post_reactions"]
        headers = [cell.value for cell in sheet[1]]
        row = [cell.value for cell in sheet[2]]
        data = dict(zip(headers, row))

        self.assertEqual(data["username"], "p001")
        self.assertEqual(data["reaction"], "like")

    def test_csv_zip_contains_selected_files(self):
        payload = build_csv_zip(self.batch, ["participants", "quality_events"])

        with zipfile.ZipFile(BytesIO(payload)) as archive:
            self.assertIn("participants.csv", archive.namelist())
            self.assertIn("quality_events.csv", archive.namelist())

    def test_all_users_excel_is_organized_by_user(self):
        payload = build_all_users_excel()

        workbook = load_workbook(BytesIO(payload))
        self.assertEqual(
            workbook.sheetnames,
            ["资料", "流程状态", "话题排序", "高分话题", "低分话题", "英文论文"],
        )

        profile_sheet = workbook["资料"]
        headers = [cell.value for cell in profile_sheet[1]]
        row = [cell.value for cell in profile_sheet[2]]
        data = dict(zip(headers, row))

        self.assertEqual(data["资料-用户名"], "p001")
        self.assertEqual(data["资料-称呼/姓名"], "参与者")
        high_sheet = workbook["高分话题"]
        high_headers = [cell.value for cell in high_sheet[1]]
        high_row = [cell.value for cell in high_sheet[2]]
        high_data = dict(zip(high_headers, high_row))
        self.assertIn("放松=5", high_data["高分话题-量表记录1"])
        self.assertIn("有帮助=6", high_data["高分话题-量表记录2"])
        self.assertNotIn("高分话题-量表记录", high_headers)
        self.assertEqual(high_data["高分话题-回复内容1"], "想法")
        self.assertNotIn("高分话题-文本记录", high_headers)

        paper_sheet = workbook["英文论文"]
        paper_headers = [cell.value for cell in paper_sheet[1]]
        paper_row = [cell.value for cell in paper_sheet[2]]
        paper_data = dict(zip(paper_headers, paper_row))
        self.assertEqual(paper_data["英文论文-内容"], "My English essay.")

        self.assertEqual(high_data["高分话题-帖子和评论互动1"], "帖子: 话题 1: 赞")
        self.assertIn("评论 #1", high_data["高分话题-帖子和评论互动2"])
        self.assertIn("你好", high_data["高分话题-AI对话1"])
        self.assertIn("回应", high_data["高分话题-AI对话2"])

    def test_all_users_csv_is_organized_by_user(self):
        payload = build_all_users_csv()

        text = payload.decode("utf-8-sig")
        rows = list(csv.DictReader(StringIO(text)))

        self.assertEqual(rows[0]["资料-用户名"], "p001")
        self.assertEqual(rows[0]["资料-称呼/姓名"], "参与者")
        self.assertIn("流程状态-当前步骤", rows[0])
        self.assertIn("话题排序", rows[0])
        self.assertIn("放松=5", rows[0]["高分话题-量表记录1"])
        self.assertIn("有帮助=6", rows[0]["高分话题-量表记录2"])
        self.assertNotIn("高分话题-量表记录", rows[0])
        self.assertEqual(rows[0]["高分话题-回复内容1"], "想法")
        self.assertNotIn("高分话题-文本记录", rows[0])
        self.assertEqual(rows[0]["英文论文-内容"], "My English essay.")
        self.assertIn("赞", rows[0]["高分话题-帖子和评论互动1"])
        self.assertIn("踩", rows[0]["高分话题-帖子和评论互动2"])
        self.assertIn("你好", rows[0]["高分话题-AI对话1"])
        self.assertIn("回应", rows[0]["高分话题-AI对话2"])

    def test_research_admin_one_click_export_downloads_all_users_csv(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        self.client.force_login(admin_user)

        response = self.client.get(reverse("research_admin_export_all"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv; charset=utf-8")
        self.assertIn("all-users-research-data", response["Content-Disposition"])
        self.assertIn(".csv", response["Content-Disposition"])
