import csv
from io import BytesIO, StringIO
from pathlib import Path

from django.conf import settings
from django.contrib import admin
from django.contrib.auth.models import Group, User
from django.core.management import call_command
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from apps.survey.models import CommentReaction, ConversationMessage, ScaleResponse, SurveySession, TextResponse
from apps.accounts.models import ParticipantProfile

from .models import AIMode, APIKey, ExperimentBatch, LLMProvider, RatingScaleConfig, ScaleItem, Topic, TopicComment


class ExperimentModelTests(TestCase):
    def test_batch_defaults_match_design(self):
        batch = ExperimentBatch.objects.create(name="批次 A")

        self.assertEqual(batch.ai_chat_minutes, 5)
        self.assertEqual(batch.english_paper_duration_hours, 24)
        self.assertIn("English", batch.english_paper_prompt)
        self.assertEqual(batch.ai_neutrality, ExperimentBatch.NEUTRALITY_MODERATE)
        self.assertEqual(batch.topic_selection_strategy, ExperimentBatch.TOPIC_STRATEGY_HIGHEST_LOWEST)
        self.assertEqual(batch.round_order_strategy, ExperimentBatch.ROUND_RANDOM)

    def test_comment_presentation_is_generated(self):
        batch = ExperimentBatch.objects.create(name="批次 A")
        topic = Topic.objects.create(batch=batch, title_zh="话题", position=1)

        comment = TopicComment.objects.create(topic=topic, body_zh="这是评论", position=1)

        self.assertEqual(comment.auto_author_name, "")
        self.assertTrue(comment.avatar_seed)
        self.assertTrue(comment.relative_time)

    def test_seed_defaults_is_idempotent(self):
        call_command("seed_defaults")
        call_command("seed_defaults")

        batch = ExperimentBatch.objects.get(name__contains="Demo Batch")
        self.assertIn("我们正在进行一项关于中国青少年的研究", batch.intro_zh)
        self.assertEqual(Topic.objects.filter(batch=batch).count(), 10)
        self.assertEqual(AIMode.objects.filter(batch=batch).count(), 3)
        self.assertEqual(ScaleItem.objects.filter(batch=batch).count(), 4)

    def test_research_console_is_staff_only_and_links_editable_content(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        batch = ExperimentBatch.objects.create(name="批次 A")
        Topic.objects.create(batch=batch, title_zh="话题 A", position=1)
        AIMode.objects.create(batch=batch, name_zh="整理信息", prompt_zh="请整理信息。")
        self.client.force_login(admin_user)

        response = self.client.get(reverse("research_admin_dashboard"))

        self.assertContains(response, "研究管理台")
        self.assertContains(response, "说明文字")
        self.assertContains(response, "话题")
        self.assertContains(response, "Prompt 设置")
        self.assertContains(response, "模型和 API")
        self.assertContains(response, reverse("research_admin_copy"))
        self.assertNotContains(response, reverse("research_admin_bulk_register"))
        self.assertNotContains(response, reverse("research_admin_export_all"))
        self.assertContains(response, 'class="research-action"', count=5)
        self.assertNotContains(response, 'class="research-action research-action-users"')
        self.assertContains(response, 'class="admin-breadcrumb-bar"')
        self.assertContains(response, 'class="admin-quick-nav"')
        self.assertNotContains(response, 'class="admin-breadcrumb-trail"')

    def test_admin_index_redirects_to_research_console(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        ExperimentBatch.objects.create(name="批次 A")
        self.client.force_login(admin_user)

        response = self.client.get(reverse("admin:index"))

        self.assertRedirects(response, reverse("research_admin_dashboard"))

    def test_admin_login_uses_polished_card_and_inline_errors(self):
        response = self.client.post(reverse("admin:login"), {"username": "missing", "password": "wrong"})

        self.assertContains(response, "admin-login-page")
        self.assertContains(response, "admin-login-card")
        self.assertContains(response, "admin-login-message")
        self.assertNotContains(response, 'class="errornote"')

    def test_copy_settings_uses_custom_form_without_repeated_title(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        ExperimentBatch.objects.create(name="批次 A", intro_zh="请按真实想法排序。")
        self.client.force_login(admin_user)

        response = self.client.get(reverse("research_admin_copy"))

        self.assertContains(response, "<h1>说明文字</h1>", html=True)
        self.assertContains(response, "第一步说明")
        self.assertContains(response, "英文论文要求")
        self.assertContains(response, "英文论文时长")
        self.assertContains(response, "返回后台主页面")
        self.assertNotContains(response, "<p><label", html=False)

    def test_topic_and_ai_mode_admin_hide_batch_concept(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        batch = ExperimentBatch.objects.create(name="批次 A")
        Topic.objects.create(batch=batch, title_zh="话题 A", position=1)
        AIMode.objects.create(batch=batch, name_zh="整理信息", prompt_zh="请整理信息。")
        self.client.force_login(admin_user)

        topic_list = self.client.get(reverse("admin:experiments_topic_changelist"))
        topic_add = self.client.get(reverse("admin:experiments_topic_add"))
        ai_mode_list = self.client.get(reverse("admin:experiments_aimode_changelist"))
        ai_mode_add = self.client.get(reverse("admin:experiments_aimode_add"))

        for response in [topic_list, topic_add, ai_mode_list, ai_mode_add]:
            self.assertNotContains(response, "批次")
            self.assertNotContains(response, "batch")

    def test_model_admin_pages_link_back_to_research_console(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        batch = ExperimentBatch.objects.create(name="Batch A")
        topic = Topic.objects.create(batch=batch, title_zh="Topic A", position=1)
        mode = AIMode.objects.create(batch=batch, name_zh="Mode A", prompt_zh="Prompt A")
        self.client.force_login(admin_user)

        responses = [
            self.client.get(reverse("admin:experiments_topic_changelist")),
            self.client.get(reverse("admin:experiments_topic_change", args=[topic.pk])),
            self.client.get(reverse("admin:experiments_aimode_changelist")),
            self.client.get(reverse("admin:experiments_aimode_change", args=[mode.pk])),
        ]

        for response in responses:
            self.assertContains(response, f'href="{reverse("research_admin_dashboard")}"')
            self.assertContains(response, "返回后台主页面")

    def test_model_admin_changelists_use_simplified_custom_shell(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        batch = ExperimentBatch.objects.create(name="批次 A")
        Topic.objects.create(batch=batch, title_zh="话题 A", position=1)
        self.client.force_login(admin_user)

        response = self.client.get(reverse("admin:experiments_topic_changelist"))

        self.assertContains(response, "#changelist")
        self.assertContains(response, "返回后台主页面")
        self.assertNotContains(response, 'name="action"')
        self.assertNotContains(response, 'id="changelist-filter"')

    def test_admin_pages_have_global_quick_navigation(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        ExperimentBatch.objects.create(name="批次 A")
        self.client.force_login(admin_user)

        response = self.client.get(reverse("admin:experiments_llmprovider_changelist"))

        self.assertContains(response, 'class="admin-quick-nav"')
        self.assertContains(response, reverse("research_admin_copy"))
        self.assertContains(response, reverse("admin:experiments_topic_changelist"))
        self.assertContains(response, reverse("research_admin_users"))
        self.assertContains(response, reverse("admin:experiments_llmprovider_changelist"))
        self.assertContains(response, reverse("admin:experiments_aimode_changelist"))

    def test_admin_quick_navigation_wraps_below_header_tools(self):
        stylesheet = (Path(settings.BASE_DIR) / "templates" / "admin" / "base_site.html").read_text(encoding="utf-8")

        self.assertIn(".admin-breadcrumb-bar {", stylesheet)
        self.assertIn("display: flex", stylesheet)
        self.assertIn("justify-content: flex-start", stylesheet)
        self.assertIn("background: #214f5f", stylesheet)

    def test_llm_provider_form_only_requires_url_model_and_api_key(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        ExperimentBatch.objects.create(name="批次 A")
        self.client.force_login(admin_user)

        response = self.client.get(reverse("admin:experiments_llmprovider_add"))

        self.assertContains(response, "返回模型和 API")
        self.assertContains(response, 'name="base_url"')
        self.assertContains(response, 'name="api_keys-0-api_key"')
        self.assertContains(response, 'name="api_keys-0-model_name"')
        self.assertContains(response, "每张卡片填写一个 API Key")
        self.assertNotContains(response, 'name="name"')
        self.assertNotContains(response, 'name="priority"')
        self.assertNotContains(response, 'name="is_active"')
        self.assertNotContains(response, "推荐模型")
        self.assertNotContains(response, "时间信息")

    def test_llm_provider_change_form_hides_object_subtitle(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        provider = LLMProvider.objects.create(
            name="deepseek",
            model_name="deepseek-r1",
            base_url="https://api.example.com/v1",
            priority=1,
        )
        self.client.force_login(admin_user)

        response = self.client.get(reverse("admin:experiments_llmprovider_change", args=[provider.pk]))

        self.assertContains(response, "修改 LLM 提供商")
        self.assertNotContains(response, "deepseek (deepseek-r1)")

    def test_llm_provider_add_saves_with_only_url_model_and_api_key(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        ExperimentBatch.objects.create(name="批次 A")
        self.client.force_login(admin_user)

        response = self.client.post(
            reverse("admin:experiments_llmprovider_add"),
            {
                "base_url": "https://api.example.com/v1",
                "api_keys-TOTAL_FORMS": "2",
                "api_keys-INITIAL_FORMS": "0",
                "api_keys-MIN_NUM_FORMS": "1",
                "api_keys-MAX_NUM_FORMS": "1000",
                "api_keys-0-api_key": "sk-example-key",
                "api_keys-0-model_name": "example-model",
                "api_keys-0-id": "",
                "api_keys-0-provider": "",
                "api_keys-1-model_name": "",
                "api_keys-1-api_key": "",
                "api_keys-1-id": "",
                "api_keys-1-provider": "",
                "_save": "保存",
            },
        )

        self.assertRedirects(response, reverse("admin:experiments_llmprovider_changelist"))
        provider = LLMProvider.objects.get(model_name="example-model")
        self.assertEqual(provider.name, "api.example.com")
        self.assertEqual(provider.base_url, "https://api.example.com/v1")
        self.assertTrue(provider.is_active)
        key = provider.api_keys.get()
        self.assertEqual(key.api_key, "sk-example-key")
        self.assertEqual(key.model_names(), ["example-model"])

    def test_llm_provider_add_supports_multiple_models_for_one_api_key(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        ExperimentBatch.objects.create(name="批次 A")
        self.client.force_login(admin_user)

        response = self.client.post(
            reverse("admin:experiments_llmprovider_add"),
            {
                "base_url": "https://api.example.com/v1",
                "api_keys-TOTAL_FORMS": "2",
                "api_keys-INITIAL_FORMS": "0",
                "api_keys-MIN_NUM_FORMS": "1",
                "api_keys-MAX_NUM_FORMS": "1000",
                "api_keys-0-api_key": "sk-a",
                "api_keys-0-model_name": "model-a\nmodel-b, model-c",
                "api_keys-0-id": "",
                "api_keys-0-provider": "",
                "api_keys-1-model_name": "",
                "api_keys-1-api_key": "",
                "api_keys-1-id": "",
                "api_keys-1-provider": "",
                "_save": "保存",
            },
        )

        self.assertRedirects(response, reverse("admin:experiments_llmprovider_changelist"))
        provider = LLMProvider.objects.get(base_url="https://api.example.com/v1")
        self.assertEqual(provider.model_name, "model-a")
        key = provider.api_keys.get()
        self.assertEqual(key.api_key, "sk-a")
        self.assertEqual(key.model_names(), ["model-a", "model-b", "model-c"])

    def test_llm_provider_changelist_shows_clear_unified_provider_table(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        ExperimentBatch.objects.create(name="批次 A")
        provider = LLMProvider.objects.create(
            name="OpenAI 主线路",
            model_name="gpt-5-mini",
            base_url="https://api.openai.com/v1",
            priority=1,
        )
        APIKey.objects.create(provider=provider, name="main", api_key="sk-test-secret-abcd", is_active=True)
        self.client.force_login(admin_user)

        response = self.client.get(reverse("admin:experiments_llmprovider_changelist"))

        self.assertContains(response, "llm-admin-page")
        self.assertContains(response, "llm-admin-card")
        self.assertContains(response, "调用顺序")
        self.assertContains(response, "https://api.openai.com/v1")
        self.assertContains(response, "gpt-5-mini")
        self.assertContains(response, "sk-test-...abcd")
        self.assertContains(response, "上移")
        self.assertContains(response, "下移")
        self.assertContains(response, "停用")
        self.assertNotContains(response, "sk-test-secret-abcd")
        self.assertNotContains(response, "更改系统配置")
        self.assertNotContains(response, 'id="changelist-search"')
        self.assertNotContains(response, 'class="object-tools"')

    def test_llm_provider_changelist_has_polished_empty_state(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        ExperimentBatch.objects.create(name="批次 A")
        self.client.force_login(admin_user)

        response = self.client.get(reverse("admin:experiments_llmprovider_changelist"))

        self.assertContains(response, "还没有配置供应商")
        self.assertContains(response, "新增供应商")
        self.assertContains(response, "填写 URL")
        self.assertContains(response, "填写模型")
        self.assertContains(response, "填写 API Key")
        self.assertContains(response, "llm-empty-state")
        self.assertNotContains(response, 'id="changelist-search"')
        self.assertNotContains(response, 'class="object-tools"')
        self.assertNotContains(response, "0 LLM 供应商")

    def test_llm_provider_order_actions_change_priority(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        first = LLMProvider.objects.create(name="A", model_name="model-a", base_url="https://a.example/v1", priority=1)
        second = LLMProvider.objects.create(name="B", model_name="model-b", base_url="https://b.example/v1", priority=2)
        self.client.force_login(admin_user)

        response = self.client.post(reverse("admin:experiments_llmprovider_move", args=[second.pk, "up"]))

        self.assertRedirects(response, reverse("admin:experiments_llmprovider_changelist"))
        first.refresh_from_db()
        second.refresh_from_db()
        self.assertLess(second.priority, first.priority)

    def test_admin_header_hides_native_user_tools(self):
        stylesheet = (Path(settings.BASE_DIR) / "templates" / "admin" / "base_site.html").read_text(encoding="utf-8")

        self.assertIn("#user-tools {", stylesheet)
        self.assertIn("display: none", stylesheet)
        self.assertIn(".theme-toggle", stylesheet)

    def test_admin_navigation_buttons_keep_readable_text_color(self):
        stylesheet = (Path(settings.BASE_DIR) / "templates" / "admin" / "base_site.html").read_text(encoding="utf-8")

        self.assertIn(".admin-quick-nav a:visited", stylesheet)
        self.assertIn("color: #0a73b8 !important", stylesheet)

    def test_topic_change_form_only_keeps_dashboard_and_changelist_buttons(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        batch = ExperimentBatch.objects.create(name="批次 A")
        topic = Topic.objects.create(batch=batch, title_zh="话题 A", position=1)
        self.client.force_login(admin_user)

        response = self.client.get(reverse("admin:experiments_topic_change", args=[topic.pk]))

        self.assertContains(response, "返回后台主页面")
        self.assertContains(response, "选择 话题 来修改")
        self.assertNotContains(response, "历史")
        self.assertNotContains(response, "History")

    def test_topic_comment_inline_hides_generated_identity_fields(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        batch = ExperimentBatch.objects.create(name="批次 A")
        topic = Topic.objects.create(batch=batch, title_zh="话题 A", position=1)
        TopicComment.objects.create(
            topic=topic,
            body_zh="评论 A",
            auto_author_name="周明",
            like_count=20,
            relative_time="刚刚",
        )
        self.client.force_login(admin_user)

        response = self.client.get(reverse("admin:experiments_topic_change", args=[topic.pk]))

        self.assertContains(response, 'name="comments-0-like_count"')
        self.assertContains(response, 'name="comments-0-relative_time"')
        self.assertNotContains(response, "作者昵称")
        self.assertNotContains(response, 'name="comments-0-auto_author_name"')
        self.assertNotContains(response, "头像种子")
        self.assertNotContains(response, "avatar_seed")

    def test_bulk_register_export_downloads_excel_without_creating_users(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        ExperimentBatch.objects.create(name="批次 A", is_active=True)
        self.client.force_login(admin_user)

        response = self.client.post(
            reverse("research_admin_bulk_register"),
            {
                "initial_password": "Start12345",
                "usernames": "student001\nstudent002",
                "action": "export",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        self.assertEqual(rows[0], ("用户名", "初始密码", "登录地址", "生成时间"))
        self.assertEqual(rows[1][0], "student001")
        self.assertEqual(rows[1][1], "Start12345")
        self.assertEqual(rows[2][0], "student002")
        self.assertFalse(User.objects.filter(username="student001").exists())
        self.assertFalse(User.objects.filter(username="student002").exists())

    def test_export_all_csv_includes_comment_reactions(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        batch = ExperimentBatch.objects.create(name="批次 A", is_active=True)
        participant = User.objects.create_user("p_react", password="Start12345")
        participant.participant_profile.display_name = "参与者"
        participant.participant_profile.batch = batch
        participant.participant_profile.save(update_fields=["display_name", "batch"])
        session = SurveySession.objects.create(
            user=participant,
            batch=batch,
            batch_snapshot={},
            topic_order_snapshot=[],
        )
        round_obj = session.rounds.create(
            round_type="high",
            topic_id=11,
            material_snapshot={
                "comments": [
                    {
                        "id": 101,
                        "author": "周明",
                        "body_zh": "这个观点很有意思。",
                        "relative_time": "刚刚",
                        "like_count": 20,
                    }
                ]
            },
        )
        CommentReaction.objects.create(round=round_obj, comment_snapshot_id=101, reaction="like")
        self.client.force_login(admin_user)

        response = self.client.get(reverse("research_admin_export_all"))

        rows = list(csv.DictReader(StringIO(response.content.decode("utf-8-sig"))))
        self.assertEqual(rows[0]["用户名"], "p_react")
        self.assertIn("comment-101", rows[0]["评论互动"])
        self.assertIn("小兔", rows[0]["评论互动"])
        self.assertIn("赞", rows[0]["评论互动"])

    def test_bulk_register_confirm_creates_users_then_shows_them_in_user_data(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        batch = ExperimentBatch.objects.create(name="批次 A", is_active=True)
        self.client.force_login(admin_user)

        response = self.client.post(
            reverse("research_admin_bulk_register"),
            {
                "initial_password": "Start12345",
                "usernames": "student001\nstudent002",
                "action": "create",
            },
            follow=True,
        )

        self.assertRedirects(response, reverse("research_admin_users"))
        self.assertEqual(User.objects.get(username="student001").participant_profile.batch, batch)
        self.assertEqual(User.objects.get(username="student002").participant_profile.batch, batch)
        self.assertContains(response, "student001")
        self.assertContains(response, "student002")
        self.assertContains(response, "未开始")

    def test_bulk_register_page_uses_custom_design_without_batch_field(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        ExperimentBatch.objects.create(name="批次 A", is_active=True)
        self.client.force_login(admin_user)

        response = self.client.get(reverse("research_admin_bulk_register"))

        self.assertContains(response, "批量注册新用户")
        self.assertContains(response, "导出账号 Excel")
        self.assertContains(response, "确认创建账号")
        self.assertContains(response, "用户名列表")
        self.assertNotContains(response, "实验批次")
        self.assertNotContains(response, 'name="batch"')

    def test_bulk_register_page_can_generate_random_usernames(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        ExperimentBatch.objects.create(name="批次 A", is_active=True)
        self.client.force_login(admin_user)

        response = self.client.get(reverse("research_admin_bulk_register"))

        self.assertContains(response, "随机生成用户名")
        self.assertContains(response, 'id="random-count"')
        self.assertContains(response, 'data-target="id_usernames"')
        self.assertContains(response, "adminGeneratedUsernames")
        self.assertContains(response, "随机生成 10 个用户名")

    def test_user_data_page_lists_registered_users_before_they_start(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        batch = ExperimentBatch.objects.create(name="批次 A", is_active=True)
        participant = User.objects.create_user("fresh_user", password="Start12345")
        participant.participant_profile.batch = batch
        participant.participant_profile.save(update_fields=["batch"])
        self.client.force_login(admin_user)

        response = self.client.get(reverse("research_admin_users"))

        self.assertContains(response, "账号状态")
        self.assertContains(response, "英文论文倒计时")
        self.assertContains(response, "fresh_user")
        self.assertContains(response, "未开始")
        self.assertContains(response, reverse("research_admin_bulk_register"))
        self.assertContains(response, reverse("research_admin_export_all"))
        self.assertContains(response, reverse("research_admin_user_detail", args=[participant.pk]))

    def test_user_detail_page_shows_single_user_records_and_countdown(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        batch = ExperimentBatch.objects.create(name="批次 A", is_active=True, english_paper_duration_hours=24)
        participant = User.objects.create_user("detail_user", password="Start12345")
        participant.participant_profile.batch = batch
        participant.participant_profile.display_name = "参与者"
        participant.participant_profile.save(update_fields=["batch", "display_name"])
        session = SurveySession.objects.create(
            user=participant,
            batch=batch,
            step_started_at={SurveySession.STEP_ENGLISH_PAPER: timezone.now().isoformat()},
            topic_order_snapshot=[],
        )
        round_obj = session.rounds.create(round_type="high", topic_id=1)
        ScaleResponse.objects.create(
            round=round_obj,
            step="emotion",
            item_type="emotion",
            item_label="我现在感到放松",
            language="zh-hans",
            min_value=1,
            max_value=7,
            selected_value=5,
        )
        TextResponse.objects.create(round=round_obj, step="initial_text", final_text="我的想法", word_count=4)
        self.client.force_login(admin_user)

        response = self.client.get(reverse("research_admin_user_detail", args=[participant.pk]))

        self.assertContains(response, "detail_user")
        self.assertContains(response, "英文倒计时")
        self.assertContains(response, "高分话题")
        self.assertContains(response, "当前感受量表")
        self.assertContains(response, "范围")
        self.assertContains(response, "AI 对话前文字回答")
        self.assertContains(response, "我的想法")
        self.assertContains(response, "返回后台主页面")

    def test_user_data_page_has_single_and_bulk_delete_controls(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        batch = ExperimentBatch.objects.create(name="批次 A", is_active=True)
        participant = User.objects.create_user("delete_ready", password="Start12345")
        participant.participant_profile.batch = batch
        participant.participant_profile.save(update_fields=["batch"])
        self.client.force_login(admin_user)

        response = self.client.get(reverse("research_admin_users"))

        self.assertContains(response, "批量删除选中用户")
        self.assertContains(response, 'name="user_ids"')
        self.assertContains(response, reverse("research_admin_delete_user", args=[participant.pk]))
        self.assertContains(response, "删除")

    def test_single_delete_user_removes_participant_account(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        batch = ExperimentBatch.objects.create(name="批次 A", is_active=True)
        participant = User.objects.create_user("delete_one", password="Start12345")
        participant.participant_profile.batch = batch
        participant.participant_profile.save(update_fields=["batch"])
        self.client.force_login(admin_user)

        response = self.client.post(reverse("research_admin_delete_user", args=[participant.pk]), follow=True)

        self.assertRedirects(response, reverse("research_admin_users"))
        self.assertFalse(User.objects.filter(username="delete_one").exists())
        self.assertContains(response, "已删除 1 个参与者账号")

    def test_bulk_delete_users_removes_selected_participant_accounts(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        batch = ExperimentBatch.objects.create(name="批次 A", is_active=True)
        first = User.objects.create_user("delete_a", password="Start12345")
        second = User.objects.create_user("delete_b", password="Start12345")
        keep = User.objects.create_user("keep_user", password="Start12345")
        for user in [first, second, keep]:
            user.participant_profile.batch = batch
            user.participant_profile.save(update_fields=["batch"])
        self.client.force_login(admin_user)

        response = self.client.post(
            reverse("research_admin_delete_users"),
            {"user_ids": [str(first.pk), str(second.pk)]},
            follow=True,
        )

        self.assertRedirects(response, reverse("research_admin_users"))
        self.assertFalse(User.objects.filter(username="delete_a").exists())
        self.assertFalse(User.objects.filter(username="delete_b").exists())
        self.assertTrue(User.objects.filter(username="keep_user").exists())
        self.assertContains(response, "已删除 2 个参与者账号")

    def test_delete_user_does_not_remove_staff_account(self):
        admin_user = User.objects.create_superuser("admin", "admin@example.com", "pass")
        batch = ExperimentBatch.objects.create(name="批次 A", is_active=True)
        staff = User.objects.create_user("staff_in_batch", password="Start12345", is_staff=True)
        staff.participant_profile.batch = batch
        staff.participant_profile.save(update_fields=["batch"])
        self.client.force_login(admin_user)

        response = self.client.post(reverse("research_admin_delete_user", args=[staff.pk]), follow=True)

        self.assertTrue(User.objects.filter(username="staff_in_batch").exists())
        self.assertContains(response, "没有可删除的参与者账号")

    def test_raw_models_are_hidden_from_admin_registry(self):
        for model in [
            ExperimentBatch,
            ScaleItem,
            RatingScaleConfig,
            SurveySession,
            ScaleResponse,
            TextResponse,
            ConversationMessage,
            ParticipantProfile,
            User,
            Group,
        ]:
            self.assertNotIn(model, admin.site._registry)

        self.assertIn(Topic, admin.site._registry)
        self.assertIn(AIMode, admin.site._registry)
