from io import BytesIO

from django import forms
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.accounts.models import ParticipantProfile
from apps.exports.services import build_all_users_csv
from apps.survey.models import CommentReaction, ConversationMessage, EnglishPaperDraft, EnglishPaperResponse, PostReaction, QualityEvent, ScaleResponse, SurveySession, TextResponse, TopicRound

from .models import AIMode, EnglishPaperConfig, ExperimentBatch, ScaleItem, Topic
from .ui_copy import UI_COPY_FIELDS, group_fields_by_step


ROUND_TYPE_LABELS = {
    "high": "高分话题",
    "low": "低分话题",
}

SESSION_STEP_LABELS = {
    "topic_order": "话题排序",
    "round": "话题答题中",
    "english_paper": "英文论文写作",
    "done": "已完成",
}

ROUND_STEP_LABELS = {
    "post": "阅读帖子与评论",
    "emotion": "当前感受量表",
    "stance_before": "AI 对话前观点量表",
    "initial_text": "AI 对话前文字回答",
    "mode": "选择 AI 对话方式",
    "chat": "AI 对话",
    "ai_eval": "AI 回复评价量表",
    "stance_after": "AI 对话后观点量表",
    "final_text": "AI 对话后文字回答",
}

ROUND_STEP_HINTS = {
    "emotion": "记录用户看到帖子后的情绪和思考状态。",
    "stance_before": "记录用户与 AI 对话前对话题的立场。",
    "initial_text": "用户在 AI 对话前写下的开放式想法。",
    "ai_eval": "用户对 AI 回复质量、中立性等方面的评价。",
    "stance_after": "记录用户与 AI 对话后对话题的立场。",
    "final_text": "用户在 AI 对话后写下的最终想法。",
}

REACTION_LABELS = {
    "like": "点赞",
    "dislike": "点踩",
    "none": "未选择",
}

ROLE_LABELS = {
    "participant": "用户",
    "assistant": "AI",
    "system": "系统",
}

INPUT_METHOD_LABELS = {
    "keyboard": "键盘输入",
    "voice": "语音输入",
}


def _display(value, fallback="-"):
    return value if value not in (None, "") else fallback


def _round_label(round_obj):
    return ROUND_TYPE_LABELS.get(round_obj.round_type, round_obj.round_type)


def _round_topic_title(round_obj, topic_titles):
    snapshot = round_obj.material_snapshot or {}
    return (
        snapshot.get("title_zh")
        or snapshot.get("title_en")
        or topic_titles.get(round_obj.topic_id)
        or f"话题 #{round_obj.topic_id}"
    )


def _step_label(step):
    return ROUND_STEP_LABELS.get(step, step)


def _snapshot_topic_title(snapshot):
    return snapshot.get("title_zh") or snapshot.get("title_en") or f"话题 #{snapshot.get('id')}"


class PageCopyForm(forms.ModelForm):
    class Meta:
        model = ExperimentBatch
        fields = ["intro_zh", "english_paper_prompt", "english_paper_duration_hours"]
        labels = {
            "intro_zh": "说明内容",
            "english_paper_prompt": "英文论文要求说明",
            "english_paper_duration_hours": "英文论文时长",
        }
        widgets = {
            "intro_zh": forms.Textarea(
                attrs={
                    "rows": 8,
                    "class": "copy-textarea",
                    "placeholder": "请按你的真实想法对以下话题排序。提交后不可返回修改。",
                }
            ),
            "english_paper_prompt": forms.Textarea(
                attrs={
                    "rows": 6,
                    "class": "copy-textarea",
                    "placeholder": "Write an English argumentative essay based on the discussion you completed.",
                }
            ),
        }


class BulkRegisterForm(forms.Form):
    batch = forms.ModelChoiceField(
        queryset=ExperimentBatch.objects.order_by("id"),
        label="加入批次",
        help_text="新创建的用户将归属到选定的批次。",
        widget=forms.Select(attrs={"class": "bulk-input"}),
    )
    initial_password = forms.CharField(
        label="初始密码",
        min_length=5,
        widget=forms.PasswordInput(
            attrs={
                "class": "bulk-input",
                "autocomplete": "new-password",
                "placeholder": "至少 5 位",
            }
        ),
    )
    usernames = forms.CharField(
        label="用户名列表",
        help_text="每行一个用户名。可先导出 Excel 留档，再确认创建账号。",
        widget=forms.Textarea(
            attrs={
                "rows": 10,
                "class": "bulk-textarea",
                "placeholder": "student001\nstudent002\nstudent003",
            }
        ),
    )

    def clean_usernames(self):
        raw = self.cleaned_data["usernames"]
        usernames = [line.strip() for line in raw.splitlines() if line.strip()]
        if not usernames:
            raise forms.ValidationError("请至少输入一个用户名。")
        duplicates = sorted({name for name in usernames if usernames.count(name) > 1})
        if duplicates:
            raise forms.ValidationError(f"用户名重复：{', '.join(duplicates)}")
        existing = set(User.objects.filter(username__in=usernames).values_list("username", flat=True))
        if existing:
            raise forms.ValidationError(f"用户名已存在：{', '.join(sorted(existing))}")
        return usernames


def default_batch():
    batch = ExperimentBatch.objects.filter(is_active=True).order_by("id").first()
    if batch:
        return batch
    batch = ExperimentBatch.objects.order_by("id").first()
    if batch:
        return batch
    return ExperimentBatch.objects.create(name="默认实验", is_active=True)


SELECTED_BATCH_SESSION_KEY = "selected_admin_batch_id"


def selected_batch(request):
    batch_id = request.session.get(SELECTED_BATCH_SESSION_KEY)
    if batch_id:
        batch = ExperimentBatch.objects.filter(pk=batch_id).first()
        if batch:
            return batch
    batch = default_batch()
    request.session[SELECTED_BATCH_SESSION_KEY] = batch.pk
    return batch


@staff_member_required
@require_POST
def select_batch(request):
    batch_id = request.POST.get("batch_id")
    if not batch_id:
        messages.error(request, "未选择批次。")
        return redirect("research_admin_dashboard")
    batch = ExperimentBatch.objects.filter(pk=batch_id).first()
    if not batch:
        messages.error(request, "批次不存在。")
        return redirect("research_admin_dashboard")
    request.session[SELECTED_BATCH_SESSION_KEY] = batch.pk
    messages.success(request, f"已切换到批次：{batch.name}")
    next_url = request.POST.get("next") or reverse("research_admin_dashboard")
    return redirect(next_url)


@staff_member_required
@require_POST
def set_register_batch(request):
    batch_id = request.POST.get("batch_id")
    batch = ExperimentBatch.objects.filter(pk=batch_id).first()
    if not batch:
        messages.error(request, "批次不存在。")
        return redirect("research_admin_dashboard")
    # 全局唯一 active：把其他所有 batch 设为 inactive，仅当前 batch 为 active
    ExperimentBatch.objects.exclude(pk=batch.pk).update(is_active=False)
    if not batch.is_active:
        batch.is_active = True
        batch.save(update_fields=["is_active"])
    messages.success(request, f"已设置「{batch.name}」为接收新用户注册的批次。")
    return redirect("research_admin_dashboard")


@staff_member_required
def delete_batch(request, batch_id):
    batch = get_object_or_404(ExperimentBatch, pk=batch_id)
    participants_count = batch.participants.count()
    sessions_count = batch.sessions.count()
    ai_modes_count = batch.ai_modes.count()
    scale_items_count = batch.scale_items.count()
    topics_unlink_count = batch.topics.count()

    if request.method == "POST":
        if ExperimentBatch.objects.count() <= 1:
            messages.error(request, "至少需要保留一个批次，无法删除最后一个。")
            return redirect("research_admin_dashboard")
        batch_name = batch.name
        # 解绑用户和 session（不删用户、不删答题数据）
        batch.participants.update(batch=None)
        batch.sessions.update(batch=None)
        # 解绑 topic（topic 本身保留，可能还属于其他 batch）
        batch.topics.clear()
        # 删 batch 本身（CASCADE 会自动删 AIMode/ScaleItem/RatingScaleConfig）
        batch.delete()
        # 清掉 session 里指向已删 batch 的引用
        if request.session.get(SELECTED_BATCH_SESSION_KEY) == batch_id:
            request.session.pop(SELECTED_BATCH_SESSION_KEY, None)
        messages.success(
            request,
            f"已删除批次「{batch_name}」。{participants_count} 个用户和 {sessions_count} 个会话已解绑到「无批次」，数据保留。",
        )
        return redirect("research_admin_dashboard")

    return render(
        request,
        "admin/research/delete_batch.html",
        {
            "title": f"删除批次「{batch.name}」",
            "batch": batch,
            "participants_count": participants_count,
            "sessions_count": sessions_count,
            "ai_modes_count": ai_modes_count,
            "scale_items_count": scale_items_count,
            "topics_unlink_count": topics_unlink_count,
            "is_last_batch": ExperimentBatch.objects.count() <= 1,
        },
    )



def _parse_started_at(raw_value):
    if not raw_value:
        return None
    started_at = parse_datetime(raw_value)
    if not started_at:
        return None
    if timezone.is_naive(started_at):
        started_at = timezone.make_aware(started_at, timezone.get_current_timezone())
    return started_at


def _format_duration(seconds):
    seconds = max(int(seconds or 0), 0)
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    seconds = seconds % 60
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def english_paper_countdown(session):
    if not session:
        return "-"
    started_at = _parse_started_at(session.step_started_at.get(SurveySession.STEP_ENGLISH_PAPER))
    if not started_at:
        return "未开始"
    duration_hours = (
        getattr(session.batch, "english_paper_duration_hours", None)
        or session.batch_snapshot.get("english_paper_duration_hours")
        or 24
    )
    total_seconds = int(duration_hours) * 3600
    elapsed = int((timezone.now() - started_at).total_seconds())
    return _format_duration(max(total_seconds - elapsed, 0))


@staff_member_required
def dashboard(request):
    batch = selected_batch(request)
    user_count = batch.participants.count()
    completed_count = batch.sessions.filter(completed_at__isnull=False).count()
    register_batch = ExperimentBatch.objects.filter(is_active=True).order_by("id").first()
    context = {
        "title": "研究管理台",
        "batch": batch,
        "register_batch": register_batch,
        "available_batches": ExperimentBatch.objects.order_by("id"),
        "topic_count": Topic.objects.filter(batches=batch).count(),
        "ai_mode_count": AIMode.objects.filter(batch=batch).count(),
        "user_count": user_count,
        "completed_count": completed_count,
        "copy_url": reverse("research_admin_copy"),
        "topics_url": reverse("admin:experiments_topic_changelist"),
        "ai_modes_url": reverse("admin:experiments_aimode_changelist"),
        "system_api_url": reverse("admin:experiments_systemapiconfig_changelist"),
        "llm_providers_url": reverse("admin:experiments_llmprovider_changelist"),
        "users_url": reverse("research_admin_users"),
        "select_batch_url": reverse("research_admin_select_batch"),
        "add_batch_url": reverse("admin:experiments_experimentbatch_add"),
        "delete_batch_url": reverse("research_admin_delete_batch", args=[batch.pk]),
        "set_register_batch_url": reverse("research_admin_set_register_batch"),
        "can_delete_batch": ExperimentBatch.objects.count() > 1,
    }
    return render(request, "admin/research/dashboard.html", context)


@xframe_options_sameorigin
@staff_member_required
def copy_preview(request):
    """Renders a participant-style page inline for the copy editor iframe.

    Query params:
      step: step key (e.g. consent, topic_order, post, ...)
      lang: zh or en
    Does NOT require a SurveySession; supplies fake context so templates render.
    """
    from types import SimpleNamespace
    from django.utils.safestring import mark_safe

    batch = selected_batch(request)
    step = request.GET.get("step", "consent")
    lang = "en" if request.GET.get("lang", "zh") == "en" else "zh"

    STEP_META_PREVIEW = {
        "consent": (1, "参与研究授权同意书", "consent", "survey/step_consent.html"),
        "topic_order": (1, "先排一排你最在意的话题", "sort", "survey/step_topic_order.html"),
        "post": (2, "阅读帖子与评论", "read", "survey/step_post.html"),
        "emotion": (3, "当前感受", "mood", "survey/step_scale.html"),
        "stance_before": (4, "你的观点", "stance", "survey/step_scale.html"),
        "initial_text": (5, "写下你的想法", "text", "survey/step_text.html"),
        "mode": (6, "选择对话模式", "mode", "survey/step_mode.html"),
        "chat": (7, "与人工智能对话", "chat", "survey/step_chat.html"),
        "ai_eval": (8, "对人工智能的评价", "ai_eval", "survey/step_scale.html"),
        "stance_after": (9, "再次确认你的观点", "stance", "survey/step_scale.html"),
        "final_text": (10, "写下你的新想法", "text", "survey/step_text.html"),
        "english_paper_intro": (11, "英文论文写作入口", "text", "survey/step_english_paper_intro.html"),
        "english_paper": (11, "英文论文写作", "text", "survey/step_english_paper.html"),
        "done": (12, "已完成，感谢你的参与", "done", "survey/done.html"),
    }
    if step not in STEP_META_PREVIEW:
        step = "consent"
    number, title, kind, template = STEP_META_PREVIEW[step]
    meta = {
        "number": number,
        "title": title,
        "kind": kind,
        "total": 12,
        "label": f"第 {number} 站",
        "route": range(1, 13),
    }

    context = {
        "batch": batch,
        "lang": lang,
        "step": step,
        "step_meta": meta,
        "preview_mode": True,
        "current_post_number": 1,
        "total_posts": 2,
    }

    # Per-step extra fake context
    if step == "topic_order":
        context["topic_order_intro"] = batch.intro_zh if lang == "zh" else (batch.intro_en or batch.intro_zh)
        context["topics"] = [
            {"id": 1, "title": "示例话题（仅预览，请到话题页编辑）", "statement": ""},
        ]
        context["form"] = SimpleNamespace(ordered_topic_ids=mark_safe('<input type="hidden" name="ordered_topic_ids" value="1">'))
    elif step == "post":
        context["material"] = {
            "title": "示例帖子标题（仅预览）",
            "post_body": "示例帖子正文。在话题页编辑实际内容。",
            "author": "示例作者",
            "avatar_file": "avatar1.png",
            "comments": [
                {"id": 1, "body": "示例评论 1。", "author": "评论者 A", "avatar_file": "avatar2.png", "relative_time": "1 小时前"},
                {"id": 2, "body": "示例评论 2。", "author": "评论者 B", "avatar_file": "avatar3.png", "relative_time": "刚刚"},
            ],
        }
    elif step in {"emotion", "ai_eval"}:
        # Use real ScaleItem rows so the inline editor can save back to DB.
        item_type = ScaleItem.EMOTION if step == "emotion" else ScaleItem.AI_EVAL
        real_items = list(batch.scale_items.filter(item_type=item_type).order_by("position", "id"))
        if real_items:
            for it in real_items:
                # Inject `values` range used by template
                max_v = it.max_value if step == "ai_eval" else 5
                it.values = range(it.min_value, max_v + 1)
                it.stance_role = None
            items = real_items
        else:
            if step == "emotion":
                items = [
                    SimpleNamespace(pk=f"e{i}", label_zh=name, label_en=name, min_value=1, max_value=5, left_label_zh="", right_label_zh="", values=range(1, 6), item_type="emotion", stance_role=None)
                    for i, name in enumerate(["好奇", "兴奋", "困惑"], start=1)
                ]
            else:
                items = [
                    SimpleNamespace(pk=f"a{i}", label_zh=name, label_en=name, min_value=1, max_value=7, left_label_zh="", right_label_zh="", values=range(1, 8), item_type="ai_eval", stance_role=None)
                    for i, name in enumerate(["AI 回复有帮助", "AI 回复保持中立"], start=1)
                ]
        context["items"] = items
        context["agreement_prompt"] = ""
        context["confidence_prompt"] = ""
        context["agreement_labels"] = []
        context["confidence_labels"] = []
        context["statement"] = ""
    elif step in {"stance_before", "stance_after"}:
        agreement_defaults = ["非常不同意", "不同意", "有点不同意", "有点同意", "同意", "非常同意"]
        confidence_defaults = ["完全不确定", "不确定", "有点不确定", "有点确定", "确定", "非常确定"]
        items = [
            SimpleNamespace(pk="prev-a", label_zh="同意度", min_value=1, max_value=6, left_label_zh="", right_label_zh="", values=range(1, 7), item_type="stance", stance_role="agreement"),
            SimpleNamespace(pk="prev-c", label_zh="确定度", min_value=1, max_value=6, left_label_zh="", right_label_zh="", values=range(1, 7), item_type="stance", stance_role="confidence"),
        ]
        context["items"] = items
        context["agreement_prompt"] = '你有多大程度上同意"示例观点"这个观点？'
        context["confidence_prompt"] = "你对自己上述的观点有多确定？"
        context["agreement_labels"] = [getattr(batch, f"agreement_label_{i}", "") or agreement_defaults[i-1] for i in range(1, 7)]
        context["confidence_labels"] = [getattr(batch, f"confidence_label_{i}", "") or confidence_defaults[i-1] for i in range(1, 7)]
        context["statement"] = "示例观点"
    elif step in {"initial_text", "final_text"}:
        context["form"] = SimpleNamespace(
            final_text=SimpleNamespace(value=lambda: ""),
            input_method=mark_safe('<input type="hidden" name="input_method" value="keyboard">'),
            transcribe_model=mark_safe('<input type="hidden" name="transcribe_model" value="">'),
        )
        context["material"] = None
    elif step == "mode":
        context["modes"] = [
            SimpleNamespace(pk=1, name_zh="提出不同观点", prompt_zh="温和地提出一个不同角度。"),
            SimpleNamespace(pk=2, name_zh="支持我的观点", prompt_zh="复述你的主要观点并补充支持理由。"),
            SimpleNamespace(pk=3, name_zh="总结信息", prompt_zh="整理问题相关的不同信息和解释。"),
        ]
    elif step == "chat":
        context["round"] = SimpleNamespace(pk=0)
        context["minutes"] = batch.ai_chat_minutes
        context["remaining_seconds"] = batch.ai_chat_minutes * 60
        context["chat_messages"] = [
            SimpleNamespace(role="participant", content="我觉得这个观点……", error_message="", was_interrupted=False),
            SimpleNamespace(role="assistant", content="从另一个角度来看……", error_message="", was_interrupted=False, model_name=""),
        ]
    elif step == "english_paper_intro":
        try:
            ep_cfg = batch.english_paper_config
        except EnglishPaperConfig.DoesNotExist:
            ep_cfg = None
        context["gate_title_zh"] = ep_cfg.gate_title_zh if ep_cfg else "即将进入写作模块"
        context["gate_title_en"] = ep_cfg.gate_title_en if ep_cfg else "You're About to Enter the Writing Module"
        context["gate_body_zh"] = ep_cfg.gate_body_zh if ep_cfg else '接下来你将进行英文论文写作。请放轻松，根据你对话题的理解，用英文写一篇论证性短文。计时将在你点击"进入写作"后开始。'
        context["gate_body_en"] = ep_cfg.gate_body_en if ep_cfg else 'You will now write a short argumentative essay in English. Take a deep breath and relax. The timer will start after you click "Start Writing".'
        context["gate_cta_zh"] = ep_cfg.gate_cta_zh if ep_cfg else "进入写作"
        context["gate_cta_en"] = ep_cfg.gate_cta_en if ep_cfg else "Start Writing"
    elif step == "english_paper":
        try:
            ep_cfg = batch.english_paper_config
        except EnglishPaperConfig.DoesNotExist:
            ep_cfg = None
        context["title_zh"] = ep_cfg.title_zh if ep_cfg else "英文论文写作"
        context["title_en"] = ep_cfg.title_en if ep_cfg else "English paper writing"
        context["intro_zh"] = ep_cfg.intro_zh if ep_cfg else "请在规定时间内完成英文论文写作。"
        context["intro_en"] = ep_cfg.intro_en if ep_cfg else "Please complete your English argumentative essay within the time limit."
        context["prompt"] = ep_cfg.prompt if ep_cfg else batch.english_paper_prompt
        duration_minutes = ep_cfg.duration_minutes if ep_cfg else int((batch.english_paper_duration_hours or 0.5) * 60)
        context["remaining_seconds"] = duration_minutes * 60
        context["deadline_at"] = 0
        context["draft"] = None
        context["form"] = SimpleNamespace(paper_text=SimpleNamespace(value=lambda: ""))
        context["session"] = None
    elif step == "done":
        context["outro"] = batch.outro_zh if lang == "zh" else (batch.outro_en or batch.outro_zh)
        context["request"] = None
        context["session"] = None

    return render(request, template, context)


@staff_member_required
def copy_settings(request):
    batch = selected_batch(request)
    ui_field_bases = [base for base, *_ in UI_COPY_FIELDS]

    if request.method == "POST":
        changed = []
        scale_changed = 0
        # Accept any field on ExperimentBatch ending in _zh or _en (whitelist by introspection).
        valid_field_names = {f.name for f in ExperimentBatch._meta.get_fields() if hasattr(f, "name")}
        for key in request.POST:
            if key.startswith("ep_"):
                continue  # handled separately below
            if key.startswith("scale_item_"):
                # scale_item_<id>_zh or scale_item_<id>_en
                if not (key.endswith("_zh") or key.endswith("_en")):
                    continue
                lang_suffix = key[-3:]  # "_zh" or "_en"
                try:
                    item_id = int(key[len("scale_item_"):-3])
                except ValueError:
                    continue
                try:
                    item = ScaleItem.objects.get(pk=item_id, batch=batch)
                except ScaleItem.DoesNotExist:
                    continue
                attr = "label_zh" if lang_suffix == "_zh" else "label_en"
                new_value = request.POST.get(key, "")
                if getattr(item, attr, "") != new_value:
                    setattr(item, attr, new_value)
                    item.save(update_fields=[attr])
                    scale_changed += 1
                continue
            if not (key.endswith("_zh") or key.endswith("_en")):
                continue
            if key not in valid_field_names:
                continue
            new_value = request.POST.get(key, "")
            if getattr(batch, key, "") != new_value:
                setattr(batch, key, new_value)
                changed.append(key)
        if changed:
            batch.save(update_fields=changed)

        # Save EnglishPaperConfig fields (ep_* prefixed keys)
        ep_changed = []
        try:
            ep_config = batch.english_paper_config
        except EnglishPaperConfig.DoesNotExist:
            ep_config = None
        if ep_config:
            ep_field_names = {f.name for f in EnglishPaperConfig._meta.get_fields() if hasattr(f, "name")}
            for key in request.POST:
                if not key.startswith("ep_"):
                    continue
                if not (key.endswith("_zh") or key.endswith("_en")):
                    continue
                field_name = key[3:]  # strip "ep_" prefix
                new_value = request.POST.get(key, "")
                if field_name in ep_field_names:
                    if getattr(ep_config, field_name, "") != new_value:
                        setattr(ep_config, field_name, new_value)
                        ep_changed.append(field_name)
                else:
                    # Field without lang suffix (e.g. ep_prompt_zh → prompt)
                    for suffix in ("_zh", "_en"):
                        if field_name.endswith(suffix):
                            base = field_name[: -len(suffix)]
                            if base in ep_field_names and base not in ep_changed:
                                if getattr(ep_config, base, "") != new_value:
                                    setattr(ep_config, base, new_value)
                                    ep_changed.append(base)
                            break
            if ep_changed:
                ep_config.save(update_fields=ep_changed)

        total_changed = len(changed) + len(ep_changed) + scale_changed
        if total_changed:
            messages.success(request, f"已保存 {total_changed} 处文案修改。")
        else:
            messages.info(request, "没有需要保存的改动。")
        return redirect("research_admin_copy")

    STEP_OPTIONS = []
    STEP_DEFS = [
        ("consent", 1),
        ("topic_order", 1),
        ("post", 2),
        ("emotion", 3),
        ("stance_before", 4),
        ("initial_text", 5),
        ("mode", 6),
        ("chat", 7),
        ("ai_eval", 8),
        ("stance_after", 9),
        ("final_text", 10),
        ("english_paper_intro", 11),
        ("english_paper", 11),
        ("done", 12),
    ]
    for step_key, number in STEP_DEFS:
        # Step title preference: consent uses consent_title; done uses done_title; others use step_title_<step>.
        if step_key == "consent":
            title = batch.consent_title_zh
        elif step_key == "done":
            title = batch.done_title_zh
        elif step_key == "english_paper_intro":
            try:
                title = batch.english_paper_config.gate_title_zh or "英文写作入口页"
            except EnglishPaperConfig.DoesNotExist:
                title = "英文写作入口页"
        elif step_key == "english_paper":
            try:
                title = batch.english_paper_config.title_zh or "英文论文写作"
            except EnglishPaperConfig.DoesNotExist:
                title = "英文论文写作"
        else:
            title = getattr(batch, f"step_title_{step_key}_zh", "")
        STEP_OPTIONS.append((step_key, f"{number}. {title}"))
    current_step = request.GET.get("step", "consent")
    current_lang = "en" if request.GET.get("lang") == "en" else "zh"

    context = {
        "title": "界面文案",
        "batch": batch,
        "step_options": STEP_OPTIONS,
        "current_step": current_step,
        "current_lang": current_lang,
        "preview_url_base": reverse("research_admin_copy_preview"),
    }
    return render(request, "admin/research/copy_settings.html", context)


@staff_member_required
def user_records(request):
    profiles = (
        ParticipantProfile.objects.all()
        .select_related("user", "user__survey_session", "batch")
        .prefetch_related("user__survey_session__rounds")
        .order_by("user__username")
    )
    rows = []
    for profile in profiles:
        session = getattr(profile.user, "survey_session", None)
        if session is None:
            status = "未开始"
            current_step = "-"
            round_count = 0
            started_at = None
            completed_at = None
            paper_countdown = "-"
        elif session.completed_at:
            status = "已完成"
            current_step = session.current_session_step
            round_count = session.rounds.count()
            started_at = session.started_at
            completed_at = session.completed_at
            paper_countdown = "已提交"
        else:
            status = "进行中"
            current_step = session.current_session_step
            round_count = session.rounds.count()
            started_at = session.started_at
            completed_at = None
            paper_countdown = english_paper_countdown(session)
        rows.append(
            {
                "user": profile.user,
                "display_name": profile.display_name,
                "batch_name": profile.batch.name if profile.batch else "—",
                "status": status,
                "current_step": current_step,
                "round_count": round_count,
                "started_at": started_at,
                "completed_at": completed_at,
                "paper_countdown": paper_countdown,
                "can_delete": not profile.user.is_staff and not profile.user.is_superuser,
                "is_tester": profile.is_tester,
            }
        )
    context = {
        "title": "用户数据",
        "rows": rows,
        "idle_count": sum(1 for row in rows if row["status"] == "未开始"),
        "completed_count": sum(1 for row in rows if row["status"] == "已完成"),
    }
    return render(request, "admin/research/user_records.html", context)


@staff_member_required
def user_detail(request, user_id):
    profile = get_object_or_404(
        ParticipantProfile.objects.select_related("user", "batch"),
        user_id=user_id,
    )
    batch = profile.batch
    session = getattr(profile.user, "survey_session", None)
    rounds = list(session.rounds.all()) if session else []
    topic_titles = {
        topic.pk: topic.title_zh or topic.title_en
        for topic in Topic.objects.filter(pk__in=[round_obj.topic_id for round_obj in rounds])
    }
    scale_responses = ScaleResponse.objects.filter(round__in=rounds).select_related("round").order_by("submitted_at", "id")
    text_responses = TextResponse.objects.filter(round__in=rounds).select_related("round").order_by("submitted_at", "id")
    conversation_messages = ConversationMessage.objects.filter(round__in=rounds).select_related("round").order_by("created_at", "id")
    post_reactions = PostReaction.objects.filter(round__in=rounds).select_related("round").order_by("submitted_at", "id")
    comment_reactions = CommentReaction.objects.filter(round__in=rounds).select_related("round").order_by("submitted_at", "id")
    scale_rows = [
        {
            "round_id": item.round_id,
            "round_label": _round_label(item.round),
            "topic_title": _round_topic_title(item.round, topic_titles),
            "step_label": _step_label(item.step),
            "step_hint": ROUND_STEP_HINTS.get(item.step, ""),
            "item_label": item.item_label,
            "value": item.selected_value,
            "range": f"{item.min_value}-{item.max_value}",
            "submitted_at": item.submitted_at,
        }
        for item in scale_responses
    ]
    text_rows = [
        {
            "round_id": item.round_id,
            "round_label": _round_label(item.round),
            "topic_title": _round_topic_title(item.round, topic_titles),
            "step_label": _step_label(item.step),
            "step_hint": ROUND_STEP_HINTS.get(item.step, ""),
            "word_count": item.word_count,
            "was_edited": item.was_edited,
            "content": item.final_text,
            "submitted_at": item.submitted_at,
        }
        for item in text_responses
    ]
    interaction_rows = [
        {
            "round_id": item.round_id,
            "type": "帖子",
            "round_label": _round_label(item.round),
            "topic_title": _round_topic_title(item.round, topic_titles),
            "target": _round_topic_title(item.round, topic_titles),
            "reaction": REACTION_LABELS.get(item.reaction, item.reaction),
            "submitted_at": item.submitted_at,
        }
        for item in post_reactions
    ] + [
        {
            "round_id": item.round_id,
            "type": "评论",
            "round_label": _round_label(item.round),
            "topic_title": _round_topic_title(item.round, topic_titles),
            "target": f"评论 #{item.comment_snapshot_id}",
            "reaction": REACTION_LABELS.get(item.reaction, item.reaction),
            "submitted_at": item.submitted_at,
        }
        for item in comment_reactions
    ]
    conversation_rows = [
        {
            "round_id": item.round_id,
            "round_label": _round_label(item.round),
            "topic_title": _round_topic_title(item.round, topic_titles),
            "role": ROLE_LABELS.get(item.role, item.get_role_display()),
            "content": item.content,
            "model_name": item.model_name,
            "error_message": item.error_message,
            "was_interrupted": item.was_interrupted,
            "created_at": item.created_at,
        }
        for item in conversation_messages
    ]
    snapshot_by_id = {}
    if session:
        snapshot_by_id = {int(item.get("id")): item for item in session.topic_order_snapshot if item.get("id") is not None}
    topic_order_rows = []
    if session:
        for index, topic_id in enumerate(session.submitted_topic_order or [], start=1):
            topic_id = int(topic_id)
            snapshot = snapshot_by_id.get(topic_id, {"id": topic_id})
            marker = ""
            if topic_id == session.selected_high_topic_id:
                marker = "高分话题"
            elif topic_id == session.selected_low_topic_id:
                marker = "低分话题"
            topic_order_rows.append(
                {
                    "rank": index,
                    "title": _snapshot_topic_title(snapshot),
                    "marker": marker,
                }
            )
    round_order = {"high": 0, "low": 1}
    ordered_rounds = sorted(rounds, key=lambda round_obj: (round_order.get(round_obj.round_type, 99), round_obj.pk))
    round_sections = [
        {
            "round": round_obj,
            "label": _round_label(round_obj),
            "topic_title": _round_topic_title(round_obj, topic_titles),
            "scale_rows": [row for row in scale_rows if row["round_id"] == round_obj.pk],
            "text_rows": [row for row in text_rows if row["round_id"] == round_obj.pk],
            "interaction_rows": [row for row in interaction_rows if row["round_id"] == round_obj.pk],
            "conversation_rows": [row for row in conversation_rows if row["round_id"] == round_obj.pk],
        }
        for round_obj in ordered_rounds
    ]
    context = {
        "title": f"用户数据 - {profile.user.username}",
        "profile": profile,
        "session": session,
        "rounds": rounds,
        "current_step_label": SESSION_STEP_LABELS.get(session.current_session_step, session.current_session_step) if session else "-",
        "paper_countdown": english_paper_countdown(session),
        "topic_order_rows": topic_order_rows,
        "round_sections": round_sections,
        "scale_rows": scale_rows,
        "text_rows": text_rows,
        "interaction_rows": interaction_rows,
        "conversation_rows": conversation_rows,
        "english_paper": EnglishPaperResponse.objects.filter(session=session).first() if session else None,
    }
    return render(request, "admin/research/user_detail.html", context)


def deletable_participants(request):
    return User.objects.filter(
        is_staff=False,
        is_superuser=False,
    )


@staff_member_required
@require_POST
def delete_user(request, user_id):
    queryset = deletable_participants(request).filter(pk=user_id)
    user_count = queryset.count()
    queryset.delete()
    if user_count:
        messages.success(request, "已删除 1 个参与者账号。")
    else:
        messages.warning(request, "没有可删除的参与者账号。")
    return redirect("research_admin_users")


@staff_member_required
@require_POST
def delete_users(request):
    user_ids = request.POST.getlist("user_ids")
    queryset = deletable_participants(request).filter(pk__in=user_ids)
    user_count = queryset.count()
    queryset.delete()
    if user_count:
        messages.success(request, f"已删除 {user_count} 个参与者账号。")
    else:
        messages.warning(request, "没有可删除的参与者账号。")
    return redirect("research_admin_users")


@staff_member_required
@require_POST
def reset_test_user(request, user_id):
    user = get_object_or_404(User, pk=user_id, is_staff=False, is_superuser=False)
    if not getattr(user.participant_profile, "is_tester", False):
        messages.error(request, "只有测试账号才可以重置。")
        return redirect("research_admin_users")
    session = getattr(user, "survey_session", None)
    if session:
        session.delete()
    QualityEvent.objects.filter(user=user).delete()
    messages.success(request, f"已重置 {user.username} 的所有答题记录，可重新开始测试。")
    return redirect("research_admin_users")


@staff_member_required
def bulk_register(request):
    current = selected_batch(request)
    if request.method == "POST":
        form = BulkRegisterForm(request.POST)
        if form.is_valid():
            password = form.cleaned_data["initial_password"]
            usernames = form.cleaned_data["usernames"]
            batch = form.cleaned_data["batch"]
            stamp = timezone.localtime().strftime("%Y%m%d-%H%M%S")
            if request.POST.get("action") == "export":
                payload = build_bulk_register_excel(request, usernames, password)
                response = HttpResponse(
                    payload,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                response["Content-Disposition"] = f'attachment; filename="participant-accounts-{stamp}.xlsx"'
                return response
            for username in usernames:
                user = User.objects.create_user(username=username, password=password)
                user.participant_profile.batch = batch
                user.participant_profile.save(update_fields=["batch"])
            messages.success(request, f"已创建 {len(usernames)} 个参与者账号到批次「{batch.name}」。")
            return redirect("research_admin_users")
    else:
        form = BulkRegisterForm(initial={"batch": current})
    return render(request, "admin/research/bulk_register.html", {"title": "批量注册新用户", "form": form, "batch": current})


def build_bulk_register_excel(request, usernames, password):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "参与者账号"
    headers = ["用户名", "初始密码", "登录地址", "生成时间"]
    sheet.append(headers)
    login_url = request.build_absolute_uri(reverse("login"))
    created_at = timezone.localtime().strftime("%Y-%m-%d %H:%M:%S")
    for username in usernames:
        sheet.append([username, password, login_url, created_at])

    header_fill = PatternFill("solid", fgColor="DFF7EE")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="172033")
        cell.fill = header_fill
    widths = [22, 20, 42, 22]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


@staff_member_required
def export_all(request):
    stamp = timezone.localtime().strftime("%Y%m%d-%H%M%S")
    payload = build_all_users_csv()
    response = HttpResponse(payload, content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="all-users-research-data-{stamp}.csv"'
    return response
